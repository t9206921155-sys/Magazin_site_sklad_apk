"""Импорт товаров из любых источников: CSV/Excel, YML-фиды (Яндекс.Маркет, поставщики), JSON API, 1С.
Все форматы приводятся к общему виду {code, name, price, description, category, photo_url|photo_base64}.
"""
import base64
import csv
import io
import logging
import os
import uuid
import xml.etree.ElementTree as ET

import requests
from PIL import Image

import config

log = logging.getLogger("shop.import")

PHOTOS_DIR = os.path.join(config.WEBAPP_DIR, "img", "products")
os.makedirs(PHOTOS_DIR, exist_ok=True)


def save_photo_data(data: str) -> str:
    """Сохраняет фото: URL -> скачивание, base64/dataURL -> декодирование. Возвращает путь или None."""
    data = (data or "").strip()
    if not data:
        return None
    try:
        if data.startswith("data:") and "base64," in data:
            raw = base64.b64decode(data.split("base64,", 1)[1])
        elif data.startswith(("http://", "https://")):
            r = requests.get(data, timeout=20, headers={"User-Agent": "TelegramShop/1.0"})
            r.raise_for_status()
            if "image" not in r.headers.get("Content-Type", ""):
                return None
            raw = r.content
        else:
            return None
        fname = f"import_{uuid.uuid4().hex[:10]}.jpg"
        with open(os.path.join(PHOTOS_DIR, fname), "wb") as f:
            f.write(raw)
        return f"/webapp/img/products/{fname}"
    except Exception as e:
        log.warning("фото не сохранено: %s", e)
        return None


def _clean(d) -> str:
    return " ".join(str(d or "").split())


def parse_csv(text: str) -> list:
    """CSV/Excel: столбцы code,name,price,description,category,photo_url,in_stock. Разделитель ; или ,"""
    text = text.lstrip("\ufeff")
    dialect = csv.Sniffer().sniff(text[:2000], delimiters=";,") if text[:1] not in (";", ",") else None
    reader = csv.DictReader(io.StringIO(text), delimiter=dialect.delimiter if dialect else (";" if ";" in text.splitlines()[0] else ","))
    out = []
    for row in reader:
        row = {(_clean(k).lower() if k else ""): v for k, v in row.items()}
        def col(*names):
            for n in names:
                for k, v in row.items():
                    if k == n or (n in k):
                        return v
            return None
        name = _clean(col("name", "название", "наименование", "товар"))
        if not name:
            continue
        try:
            price = float(str(col("price", "цена")).replace(",", ".").replace(" ", "") or 0)
        except ValueError:
            price = 0
        try:
            old_price = float(str(col("old_price", "oldprice", "старая_цена")).replace(",", ".").replace(" ", "") or 0)
        except ValueError:
            old_price = 0
        try:
            stock = int(float(str(col("stock", "остаток", "кол_во")).replace(" ", "") or -1))
        except ValueError:
            stock = -1
        out.append({
            "code": _clean(col("code", "артикул", "код", "vendorcode", "sku")),
            "name": name,
            "price": int(price),
            "old_price": int(old_price),
            "stock": stock,
            "description": _clean(col("description", "описание")),
            "category": _clean(col("category", "категория")),
            "photo_url": _clean(col("photo_url", "photo", "фото", "picture", "image")),
            "in_stock": str(col("in_stock", "наличие") or "true").lower() not in ("0", "false", "нет"),
        })
    return out


def parse_yml(xml_text: str) -> list:
    """YML (Yandex Market Language): фиды поставщиков и агрегаторов."""
    root = ET.fromstring(xml_text)
    shop = root.find("shop")
    if shop is None:
        return []
    categories = {c.get("id"): _clean(c.text) for c in shop.findall("categories/category")}
    out = []
    for offer in shop.findall("offers/offer"):
        name = _clean(offer.findtext("name")) or _clean(offer.findtext("model"))
        if not name:
            continue
        picture = offer.findtext("picture")
        price = float(offer.findtext("price") or 0)
        cat = categories.get(offer.findtext("categoryId"), "")
        out.append({
            "code": _clean(offer.get("id")) or _clean(offer.findtext("vendorCode")),
            "name": name,
            "price": int(price),
            "description": _clean(offer.findtext("description")),
            "category": cat,
            "photo_url": picture,
            "in_stock": str(offer.get("available", "true")).lower() != "false",
        })
    return out


def parse_json_items(payload) -> list:
    """JSON API/1С: {"products": [...]} или просто список. Поддерживает photo_base64."""
    items = payload.get("products") if isinstance(payload, dict) else payload
    if not isinstance(items, list):
        raise ValueError("Ожидается список товаров")
    out = []
    for it in items:
        if not isinstance(it, dict):
            continue
        out.append({
            "code": _clean(it.get("code") or it.get("id") or it.get("sku")),
            "name": _clean(it.get("name") or it.get("title")),
            "price": int(float(it.get("price") or 0)),
            "old_price": int(float(it.get("old_price") or 0)),
            "stock": int(float(it.get("stock") or -1)),
            "description": _clean(it.get("description")),
            "category": _clean(it.get("category")),
            "photo_url": _clean(it.get("photo") or it.get("photo_url")),
            "photo_base64": it.get("photo_base64") or "",
            "in_stock": bool(it.get("in_stock", True)),
        })
    return [i for i in out if i["name"]]


def parse_xlsx(data_bytes) -> list:
    """Импорт из Excel (.xlsx): колонки как в экспорте + фото.
    Фото берётся из колонки «фото» (URL или base64) или из встроенной картинки ячейки."""
    import base64 as b64
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [(str(h) or "").strip().lower() for h in rows[0]]

    img_by_row = {}
    for img in getattr(ws, "_images", []):
        try:
            row_idx = img.anchor._from.row
            img_by_row.setdefault(row_idx, []).append(img._data())
        except Exception:
            continue

    out = []
    for r_idx, row in enumerate(rows[1:], start=1):
        rec = {headers[i]: ("" if row[i] is None else row[i])
               for i in range(min(len(headers), len(row)))}

        def col(*names):
            for n in names:
                if n in rec and str(rec[n]).strip():
                    return rec[n]
            return None

        name = _clean(col("name", "название", "наименование", "товар"))
        if not name:
            continue

        def num(v, default=0):
            try:
                return float(str(v).replace(",", ".").replace(" ", "") or default)
            except ValueError:
                return default

        photo_url = ""
        photo_base64 = ""
        pcell = col("фото", "фото (url)", "photo", "photo_url", "картинка")
        if pcell:
            pv = str(pcell).strip()
            if pv.startswith(("http://", "https://", "/webapp/", "/site/")):
                photo_url = pv
            elif pv.startswith("data:"):
                photo_base64 = pv
        elif r_idx in img_by_row:
            photo_base64 = "data:image/jpeg;base64," + b64.b64encode(img_by_row[r_idx][0]).decode()

        out.append({
            "code": _clean(col("code", "артикул", "код", "sku", "vendorcode")),
            "name": name,
            "price": int(num(col("price", "цена"))),
            "old_price": int(num(col("old_price", "старая цена", "oldprice"))),
            "stock": int(num(col("stock", "остаток", "кол-во", "количество"), default=-1)),
            "description": _clean(col("description", "описание")),
            "category": _clean(col("category", "категория")),
            "photo_url": photo_url,
            "photo_base64": photo_base64,
            "in_stock": str(col("in_stock", "наличие") or "true").lower() not in ("0", "false", "нет"),
        })
    return out


def export_products_xlsx(store, with_images: bool = True) -> bytes:
    """Экспорт товаров в Excel (.xlsx) со встроенными миниатюрами фото."""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"
    headers = ["id", "артикул", "название", "категория", "цена", "старая цена",
               "остаток", "в наличии", "бейджи", "описание", "фото (url)"]
    if with_images:
        headers.append("фото")
    ws.append(headers)
    widths = [6, 14, 42, 16, 10, 12, 9, 10, 18, 60, 44]
    for i, wd in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    if with_images:
        ws.column_dimensions[get_column_letter(len(headers))].width = 10

    for p in store.products():
        ws.append([
            p["id"], p.get("code", ""), p["name"], p.get("category", ""),
            p["price"], p.get("old_price", 0), p.get("stock", -1),
            "да" if p.get("in_stock") else "нет",
            ",".join(p.get("badges", [])),
            p.get("description", ""),
            p.get("photo", "") if str(p.get("photo", "")).startswith(("http", "/")) else "",
        ])
        if with_images:
            row = ws.max_row
            ws.row_dimensions[row].height = 58
            try:
                photo = p.get("photo") or ""
                path = ""
                if photo.startswith("/webapp/"):
                    path = os.path.join(config.WEBAPP_DIR, photo.split("/webapp/", 1)[1])
                elif photo.startswith("/site/"):
                    path = os.path.join(config.SITE_DIR, photo.split("/site/", 1)[1])
                elif photo.startswith(("http://", "https://")):
                    r = requests.get(photo, timeout=15)
                    raw = io.BytesIO(r.content)
                else:
                    continue
                if path:
                    raw = io.BytesIO(open(path, "rb").read())
                im = Image.open(raw)
                im.thumbnail((54, 54))
                buf = io.BytesIO()
                im.convert("RGB").save(buf, "JPEG", quality=85)
                buf.seek(0)
                ximg = XLImage(buf)
                ximg.width, ximg.height = 54, 54
                ws.add_image(ximg, f"{get_column_letter(len(headers))}{row}")
            except Exception as e:
                log.warning("экспорт xlsx: фото %s: %s", p.get("id"), e)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def apply_import(store, items: list) -> dict:
    """Применяет список товаров к каталогу (upsert по коду/названию) и подтягивает фото."""
    created = updated = skipped = 0
    for it in items:
        if not it.get("name"):
            skipped += 1
            continue
        photo = save_photo_data(it.get("photo_base64") or it.get("photo_url") or "")
        data = {k: v for k, v in it.items() if k in
                ("code", "name", "price", "old_price", "stock", "description", "category", "in_stock")}
        data["photo"] = photo or None
        action, _ = store.upsert_by_code(data)
        if action == "created":
            created += 1
        else:
            updated += 1
    return {"created": created, "updated": updated, "skipped": skipped}
